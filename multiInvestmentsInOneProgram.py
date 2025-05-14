import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Function to calculate compound interest for multiple investments
def compound_interest_multiple():
    all_results = []  # Store results for all investments
    grand_total_interest = 0
    grand_total_balance = 0

    first_investment = True  # Track if this is the first investment

    while True:
        # Get user input for investment details
        startBal = float(input("\nEnter starting balance for investment): "))
        rate = float(input("Enter annual interest rate (%): ")) / 100
        numMonths = int(input("Enter number of months: "))

        results = []  # Store individual investment results
        initBal = startBal

        # Add headers only **before each investment**, not every row
        if first_investment or len(all_results) > 0:
            all_results.append({
                "Month": "Month",
                "Start Balance": "Start Balance",
                "Interest Rate (%)": "Interest Rate (%)",
                "End Balance": "End Balance",
                "Interest Earned": "Interest Earned"
            })
            first_investment = False  # Prevent additional headers

        for i in range(numMonths):
            runningBal = startBal
            startBal *= (1 + rate / 12)
            monthlyInterestEarned = startBal - runningBal

            results.append({
                "Month": i + 1,
                "Start Balance": round(runningBal, 2),
                "Interest Rate (%)": round(rate * 100, 2),
                "End Balance": round(startBal, 2),
                "Interest Earned": round(monthlyInterestEarned, 2)
            })

        # Calculate totals for this investment
        investment_total_interest = round(startBal - initBal, 2)
        grand_total_interest += investment_total_interest
        grand_total_balance += startBal

        print(f"\nInvestment Summary:")
        print(f"- Initial Balance: ${initBal:,.2f}")
        print(f"- Final Balance: ${startBal:,.2f}")
        print(f"- Interest Earned: ${investment_total_interest:,.2f}")

        #Adding a total interest earned to the spreadsheet
        results.append({
            "Total Earnings $": round(grand_total_interest, 2)
        })

        # Append investment data to overall list
        all_results.extend(results)

        # Add a blank row for spacing between investments
        all_results.append({
            "Month": "",
            "Start Balance": "",
            "Interest Rate (%)": "",
            "End Balance": "",
            "Interest Earned": ""
        })

        add_investment = input("\nWould you like to enter a new investment? (yes/no): ").strip().lower()
        if add_investment not in ("yes", "y", "Yes", "Y"):
            break  

    # Create DataFrame from all investments
    df = pd.DataFrame(all_results)

    # Save results to Excel
    excel_filename = "compound_interest.xlsx"
    df.to_excel(excel_filename, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(excel_filename)
    ws = wb.active

    # **Center-align all headers**
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].value == "Month":  # Identify header rows
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    # Save the formatted Excel file
    wb.save(excel_filename)

    # Display grand totals
    print("\nFinal Grand Totals:")
    print(f"- Total Interest Earned: ${grand_total_interest:,.2f}")
    print(f"- Grand Total Balance: ${grand_total_balance:,.2f}")

# Run the program
compound_interest_multiple()

